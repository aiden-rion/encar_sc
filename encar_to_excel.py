# encar_to_excel.py
# 엔카 목록(100대) -> vehicle(carId) -> (vehicle 응답에서 userId 추출)
# -> inspection(성능점검) 세부항목 + record(보험이력) + options(choice 옵션)
# -> 엑셀 저장 + Summary(종합) 시트 생성

import time
import json
import random
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


# =========================
# 설정
# =========================
LIST_URL = (
    "https://api.encar.com/search/car/list/pricesupply"
    "?count=true"
    "&q=(And.Hidden.N._.ServiceCopyCar.Original._.(C.CarType.A._.Manufacturer.%ED%98%84%EB%8C%80.)_.Mileage.range(0..200000).)"
    "&sr=%7CModifiedDate%7C5%7C200"
)

INSPECTION_URL = "https://api.encar.com/v1/readside/inspection/vehicle/{carId}"
VEHICLE_URL = "https://api.encar.com/v1/readside/vehicle/{carId}"
USER_URL = "https://api.encar.com/v1/readside/user/{userId}"

# ✅ 보험이력 (open)
RECORD_OPEN_URL = "https://api.encar.com/v1/readside/record/vehicle/{carId}/open?vehicleNo={vehicleNo}"

# ✅ 옵션 (choice)
OPTIONS_CHOICE_URL = "https://api.encar.com/v1/readside/vehicles/car/{carId}/options/choice"

OUTPUT_XLSX = "encar_100cars_detail.xlsx"

SLEEP_BETWEEN_CALLS_SEC = (0.15, 0.35)
MAX_RETRIES = 4
TIMEOUT_SEC = 15

DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
    "Referer": "https://www.encar.com/",
    "Origin": "https://www.encar.com",
}


# =========================
# 유틸
# =========================
def sleepy():
    time.sleep(random.uniform(*SLEEP_BETWEEN_CALLS_SEC))


def safe_get(d: Any, path: List[Any], default=None):
    cur = d
    for p in path:
        if cur is None:
            return default
        try:
            if isinstance(p, int):
                if isinstance(cur, list) and 0 <= p < len(cur):
                    cur = cur[p]
                else:
                    return default
            else:
                if isinstance(cur, dict) and p in cur:
                    cur = cur[p]
                else:
                    return default
        except Exception:
            return default
    return cur if cur is not None else default


def first_not_none(*vals):
    for v in vals:
        if v is not None and v != "":
            return v
    return None


def flatten_json(obj: Any, prefix: str = "", out: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    JSON을 1-depth dict로 평탄화
    - list는 컬럼 폭발 방지를 위해 JSON 문자열로 저장
    """
    if out is None:
        out = {}

    if isinstance(obj, dict):
        for k, v in obj.items():
            new_prefix = f"{prefix}.{k}" if prefix else str(k)
            flatten_json(v, new_prefix, out)
    elif isinstance(obj, list):
        out[prefix] = json.dumps(obj, ensure_ascii=False)
    else:
        out[prefix] = obj
    return out


class EncarClient:
    def __init__(self, headers: Dict[str, str] = None):
        self.s = requests.Session()
        self.s.headers.update(headers or DEFAULT_HEADERS)

    def get_json(self, url: str) -> Any:
        last_err = None
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                resp = self.s.get(url, timeout=TIMEOUT_SEC)
                if resp.status_code == 200:
                    return resp.json()

                if resp.status_code in (429, 403, 502, 503, 504):
                    last_err = RuntimeError(f"HTTP {resp.status_code}: {resp.text[:200]}")
                    time.sleep(0.7 * attempt)
                    continue

                resp.raise_for_status()
            except Exception as e:
                last_err = e
                time.sleep(0.7 * attempt)

        raise RuntimeError(f"GET failed after retries: {url}\nlast_err={last_err}")

    def fetch_list_100(self) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        data = self.get_json(LIST_URL)
        cars = extract_list_items(data)
        return cars, data

    def fetch_vehicle(self, car_id: str) -> Dict[str, Any]:
        return self.get_json(VEHICLE_URL.format(carId=car_id))

    def fetch_inspection(self, car_id: str) -> Dict[str, Any]:
        return self.get_json(INSPECTION_URL.format(carId=car_id))

    def fetch_user(self, user_id: str) -> Dict[str, Any]:
        return self.get_json(USER_URL.format(userId=user_id))

    # ✅ 보험이력
    def fetch_record_open(self, car_id: str, vehicle_no: str) -> Dict[str, Any]:
        return self.get_json(RECORD_OPEN_URL.format(carId=car_id, vehicleNo=vehicle_no))

    # ✅ 옵션(choice)
    def fetch_options_choice(self, car_id: str) -> Any:
        return self.get_json(OPTIONS_CHOICE_URL.format(carId=car_id))


def extract_list_items(list_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    candidates = [
        ["SearchResults"],
        ["searchResults"],
        ["items"],
        ["cars"],
        ["Data", "SearchResults"],
        ["data", "SearchResults"],
        ["data", "items"],
    ]
    for path in candidates:
        arr = safe_get(list_json, path)
        if isinstance(arr, list) and len(arr) > 0:
            return arr

    stack = [list_json]
    while stack:
        cur = stack.pop()
        if isinstance(cur, dict):
            for v in cur.values():
                if isinstance(v, list) and v and isinstance(v[0], dict):
                    return v
                stack.append(v)
        elif isinstance(cur, list):
            for v in cur:
                stack.append(v)
    return []


def pick_carid_from_list_item(item: Dict[str, Any]) -> Optional[str]:
    car_id = (
        item.get("Id")
        or item.get("id")
        or item.get("CarId")
        or item.get("carId")
        or safe_get(item, ["Vehicle", "Id"])
        or safe_get(item, ["Vehicle", "id"])
    )
    return str(car_id) if car_id is not None else None


def pick_userid_from_vehicle(vehicle_json: Dict[str, Any]) -> Optional[str]:
    uid = (
        safe_get(vehicle_json, ["contact", "userId"])
        or safe_get(vehicle_json, ["contact", "UserId"])
        or safe_get(vehicle_json, ["partnership", "dealer", "userId"])
        or safe_get(vehicle_json, ["partnership", "dealer", "UserId"])
        or vehicle_json.get("userId")
        or vehicle_json.get("UserId")
        or vehicle_json.get("sellerUserId")
        or vehicle_json.get("SellerUserId")
        or vehicle_json.get("dealerUserId")
        or vehicle_json.get("DealerUserId")
        or safe_get(vehicle_json, ["Seller", "userId"])
        or safe_get(vehicle_json, ["Seller", "UserId"])
        or safe_get(vehicle_json, ["Dealer", "userId"])
        or safe_get(vehicle_json, ["Dealer", "UserId"])
        or safe_get(vehicle_json, ["seller", "userId"])
        or safe_get(vehicle_json, ["dealer", "userId"])
    )
    return str(uid) if uid is not None else None


# ✅ record(open) 호출용 vehicleNo 추출 (응답 구조가 바뀔 수 있어 후보를 넓게 둠)
def pick_vehicle_no_from_vehicle(v: Dict[str, Any]) -> Optional[str]:
    vn = first_not_none(
        v.get("vehicleNo"),
        v.get("VehicleNo"),
        v.get("carNo"),
        v.get("CarNo"),
        safe_get(v, ["vehicleNo"]),
        safe_get(v, ["VehicleNo"]),
        safe_get(v, ["spec", "vehicleNo"]),
        safe_get(v, ["Spec", "vehicleNo"]),
        safe_get(v, ["registration", "carNo"]),
        safe_get(v, ["Registration", "carNo"]),
    )
    if vn is None:
        return None
    return str(vn)


# =========================
# vehicle 주요정보(요약용) 추출
# =========================
def extract_vehicle_summary_fields(v: Dict[str, Any]) -> Dict[str, Any]:
    title = first_not_none(
        v.get("title"), v.get("Title"),
        v.get("carName"), v.get("CarName"),
        safe_get(v, ["vehicle", "title"]),
        safe_get(v, ["Vehicle", "title"]),
        safe_get(v, ["Vehicle", "Title"]),
        safe_get(v, ["model", "title"]),
        safe_get(v, ["Model", "title"]),
        safe_get(v, ["car", "title"]),
    )

    year = first_not_none(
        v.get("year"), v.get("Year"),
        v.get("modelYear"), v.get("ModelYear"),
        safe_get(v, ["spec", "year"]),
        safe_get(v, ["Spec", "year"]),
        safe_get(v, ["Vehicle", "year"]),
        safe_get(v, ["Vehicle", "Year"]),
    )

    mileage = first_not_none(
        v.get("mileage"), v.get("Mileage"),
        v.get("km"), v.get("Km"),
        safe_get(v, ["spec", "mileage"]),
        safe_get(v, ["Spec", "mileage"]),
        safe_get(v, ["Vehicle", "mileage"]),
    )

    price = first_not_none(
        v.get("price"), v.get("Price"),
        v.get("salePrice"), v.get("SalePrice"),
        safe_get(v, ["pricing", "price"]),
        safe_get(v, ["Pricing", "price"]),
        safe_get(v, ["Vehicle", "price"]),
    )

    region = first_not_none(
        v.get("region"), v.get("Region"),
        safe_get(v, ["location", "region"]),
        safe_get(v, ["Location", "region"]),
        safe_get(v, ["dealer", "region"]),
        safe_get(v, ["Dealer", "region"]),
    )

    return {
        "vehicle.title": title,
        "vehicle.year": year,
        "vehicle.mileage": mileage,
        "vehicle.price": price,
        "vehicle.region": region,
    }


# =========================
# inspection "세부내역" 추출 + 이슈 요약
# =========================
def _type_fields(node: Dict[str, Any]) -> Tuple[Optional[str], Optional[str]]:
    t = node.get("type") if isinstance(node, dict) else None
    if isinstance(t, dict):
        return (t.get("code"), t.get("title"))
    return (None, None)


def explode_inspection_tree(car_id: str, inspection_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    vehicle_id = inspection_json.get("vehicleId")

    def walk(nodes: Any, section: str, parent_code: Optional[str], parent_title: Optional[str], depth: int, path: str):
        if not isinstance(nodes, list):
            return
        for n in nodes:
            if not isinstance(n, dict):
                continue

            code, title = _type_fields(n)

            status_type = n.get("statusType") if isinstance(n.get("statusType"), dict) else None
            status_code = status_type.get("code") if status_type else None
            status_title = status_type.get("title") if status_type else None

            row = {
                "carId": car_id,
                "vehicleId": vehicle_id,
                "section": section,
                "depth": depth,
                "path": path,
                "type.code": code,
                "type.title": title,
                "parent.code": parent_code,
                "parent.title": parent_title,
                "status.code": status_code,
                "status.title": status_title,
                "description": n.get("description"),
                "price": n.get("price"),
                "exists": n.get("exists"),
                "statusItemTypes": json.dumps(n.get("statusItemTypes"), ensure_ascii=False) if isinstance(n.get("statusItemTypes"), list) else None,
                "statusTypes": json.dumps(n.get("statusTypes"), ensure_ascii=False) if isinstance(n.get("statusTypes"), list) else None,
            }
            rows.append(row)

            children = n.get("children")
            new_path = f"{path} > {title}" if title else path
            walk(children, section, code, title, depth + 1, new_path)

    walk(inspection_json.get("inners"), "inners", None, None, 0, "inners")
    walk(inspection_json.get("outers"), "outers", None, None, 0, "outers")
    return rows


def explode_inspection_images(car_id: str, inspection_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    vehicle_id = inspection_json.get("vehicleId")
    images = inspection_json.get("images")
    if not isinstance(images, list):
        return rows
    for img in images:
        if not isinstance(img, dict):
            continue
        rows.append({
            "carId": car_id,
            "vehicleId": vehicle_id,
            "type": img.get("type"),
            "title": img.get("title"),
            "path": img.get("path"),
        })
    return rows


def explode_inspection_etcs(car_id: str, inspection_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    vehicle_id = inspection_json.get("vehicleId")

    def walk(nodes: Any, parent_title: Optional[str], depth: int, path: str):
        if not isinstance(nodes, list):
            return
        for n in nodes:
            if not isinstance(n, dict):
                continue

            code, title = _type_fields(n)

            rows.append({
                "carId": car_id,
                "vehicleId": vehicle_id,
                "depth": depth,
                "path": path,
                "type.code": code,
                "type.title": title,
                "parent.title": parent_title,
                "exists": n.get("exists"),
                "statusItemTypes": json.dumps(n.get("statusItemTypes"), ensure_ascii=False) if isinstance(n.get("statusItemTypes"), list) else None,
                "statusTypes": json.dumps(n.get("statusTypes"), ensure_ascii=False) if isinstance(n.get("statusTypes"), list) else None,
            })

            children = n.get("children")
            new_path = f"{path} > {title}" if title else path
            walk(children, title, depth + 1, new_path)

    etcs = inspection_json.get("etcs")
    if isinstance(etcs, list):
        walk(etcs, None, 0, "etcs")
    return rows


def build_inspection_issue_summary(item_rows: List[Dict[str, Any]], max_items: int = 6) -> Tuple[int, str]:
    issues = []
    for r in item_rows:
        st = r.get("status.title")
        title = r.get("type.title")
        desc = r.get("description")

        if not title:
            continue

        if st is None:
            continue
        if isinstance(st, str) and st.strip() == "정상":
            continue

        if desc:
            issues.append(f"{title}({st})-{desc}")
        else:
            issues.append(f"{title}({st})")

    issue_count = len(issues)
    issue_top = " | ".join(issues[:max_items]) if issues else ""
    return issue_count, issue_top


# =========================
# ✅ 보험이력(record) 요약 + 사고내역 explode
# =========================
def explode_record_accidents(car_id: str, record_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    accidents = record_json.get("accidents")
    if not isinstance(accidents, list):
        return rows

    for a in accidents:
        if not isinstance(a, dict):
            continue
        rows.append({
            "carId": car_id,
            "type": a.get("type"),
            "date": a.get("date"),
            "insuranceBenefit": a.get("insuranceBenefit"),
            "partCost": a.get("partCost"),
            "laborCost": a.get("laborCost"),
            "paintingCost": a.get("paintingCost"),
        })
    return rows


def build_record_summary(record_json: Optional[Dict[str, Any]], max_items: int = 3) -> Dict[str, Any]:
    if not isinstance(record_json, dict):
        return {
            "record.openData": None,
            "record.accidentCnt": None,
            "record.myAccidentCnt": None,
            "record.otherAccidentCnt": None,
            "record.ownerChangeCnt": None,
            "record.totalLossCnt": None,
            "record.floodTotalLossCnt": None,
            "record.myAccidentCost": None,
            "record.otherAccidentCost": None,
            "record.accidentTop": None,
        }

    accidents = record_json.get("accidents") if isinstance(record_json.get("accidents"), list) else []
    tops = []
    for a in accidents[:max_items]:
        if not isinstance(a, dict):
            continue
        d = a.get("date")
        t = a.get("type")
        ben = a.get("insuranceBenefit")
        tops.append(f"{d}/type{t}/benefit{ben}")

    return {
        "record.openData": record_json.get("openData"),
        "record.accidentCnt": record_json.get("accidentCnt"),
        "record.myAccidentCnt": record_json.get("myAccidentCnt"),
        "record.otherAccidentCnt": record_json.get("otherAccidentCnt"),
        "record.ownerChangeCnt": record_json.get("ownerChangeCnt"),
        "record.totalLossCnt": record_json.get("totalLossCnt"),
        "record.floodTotalLossCnt": record_json.get("floodTotalLossCnt"),
        "record.myAccidentCost": record_json.get("myAccidentCost"),
        "record.otherAccidentCost": record_json.get("otherAccidentCost"),
        "record.accidentTop": " | ".join(tops) if tops else "",
    }


# =========================
# ✅ 옵션(choice) 요약 + raw 저장
# =========================
def build_options_choice_summary(choice_json: Any, max_items: int = 15) -> Dict[str, Any]:
    """
    options/choice 응답은 보통 list 형태(선택된 옵션 목록)로 옴.
    """
    if not isinstance(choice_json, list):
        return {
            "options.count": None,
            "options.totalPrice": None,
            "options.namesTop": None,
            "options.codesTop": None,
        }

    names = []
    codes = []
    total_price = 0
    for it in choice_json:
        if not isinstance(it, dict):
            continue
        name = it.get("name") or it.get("title")
        code = it.get("code") or it.get("id")
        price = it.get("price")

        if name:
            names.append(str(name))
        if code:
            codes.append(str(code))
        if isinstance(price, (int, float)):
            total_price += price

    return {
        "options.count": len([x for x in choice_json if isinstance(x, dict)]),
        "options.totalPrice": total_price if total_price != 0 else None,
        "options.namesTop": " | ".join(names[:max_items]) if names else "",
        "options.codesTop": " | ".join(codes[:max_items]) if codes else "",
    }


def normalize_options_choice_rows(car_id: str, choice_json: Any) -> List[Dict[str, Any]]:
    """
    options_choice_raw 시트에 차량별로 row 풀어서 저장
    """
    rows: List[Dict[str, Any]] = []
    if not isinstance(choice_json, list):
        return rows
    for it in choice_json:
        if not isinstance(it, dict):
            continue
        row = {"carId": car_id}
        # 그대로 넣되, 안전하게 flatten
        row.update(flatten_json(it, prefix="option"))
        rows.append(row)
    return rows


# =========================
# 엑셀 작성
# =========================
def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 70)


def write_sheet(ws, rows: List[Dict[str, Any]], title: str):
    ws.title = title
    if not rows:
        ws.append(["NO DATA"])
        return

    keys = sorted({k for r in rows for k in r.keys()})
    ws.append(keys)

    header_font = Font(bold=True)
    for c in range(1, len(keys) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for r in rows:
        ws.append([r.get(k) for k in keys])

    ws.freeze_panes = "A2"
    autosize_columns(ws)


# =========================
# 메인
# =========================
def main():
    client = EncarClient()

    print("1) Fetch list (100 cars)...")
    list_items, _ = client.fetch_list_100()
    if not list_items:
        raise RuntimeError("목록에서 차량 배열을 찾지 못했습니다. LIST_URL 응답 구조를 확인하세요.")
    list_items = list_items[:200]

    # 원본/상세 시트용
    list_rows: List[Dict[str, Any]] = []
    vehicle_rows: List[Dict[str, Any]] = []
    inspection_rows: List[Dict[str, Any]] = []
    user_rows: List[Dict[str, Any]] = []

    # inspection 세부 시트용
    inspection_item_rows: List[Dict[str, Any]] = []
    inspection_image_rows: List[Dict[str, Any]] = []
    inspection_etc_rows: List[Dict[str, Any]] = []

    # ✅ record 시트용
    record_rows: List[Dict[str, Any]] = []
    record_accident_rows: List[Dict[str, Any]] = []

    # ✅ options 시트용
    options_choice_rows: List[Dict[str, Any]] = []

    # ✅ 종합 요약 시트용(차량당 1행)
    summary_rows: List[Dict[str, Any]] = []

    user_cache: Dict[str, Dict[str, Any]] = {}

    print(f"2) Fetch details for {len(list_items)} cars...")
    for idx, item in enumerate(list_items, start=1):
        car_id = pick_carid_from_list_item(item)
        list_rows.append(flatten_json(item, prefix="list"))

        if not car_id:
            print(f"[{idx}] carId not found in list item. skip.")
            continue

        print(f"[{idx}] carId={car_id}")

        # -----------------
        # vehicle
        # -----------------
        sleepy()
        v = None
        vehicle_summary = {}
        user_id = None
        vehicle_no = None

        try:
            v = client.fetch_vehicle(car_id)
            vflat = flatten_json(v, prefix="vehicle")
            vflat["meta.carId"] = car_id
            vehicle_rows.append(vflat)

            user_id = pick_userid_from_vehicle(v) if isinstance(v, dict) else None
            vehicle_no = pick_vehicle_no_from_vehicle(v) if isinstance(v, dict) else None
            vehicle_summary = extract_vehicle_summary_fields(v) if isinstance(v, dict) else {}
        except Exception as e:
            vehicle_rows.append({"meta.carId": car_id, "meta.error_vehicle": str(e)})

        # -----------------
        # inspection (핵심)
        # -----------------
        sleepy()
        car_item_rows_this: List[Dict[str, Any]] = []
        issue_count = 0
        issue_top = ""

        try:
            ins = client.fetch_inspection(car_id)

            # raw
            iflat = flatten_json(ins, prefix="inspection")
            iflat["meta.carId"] = car_id
            inspection_rows.append(iflat)

            # explode
            car_item_rows_this = explode_inspection_tree(car_id, ins)
            inspection_item_rows.extend(car_item_rows_this)
            inspection_image_rows.extend(explode_inspection_images(car_id, ins))
            inspection_etc_rows.extend(explode_inspection_etcs(car_id, ins))

            # issue summary
            issue_count, issue_top = build_inspection_issue_summary(car_item_rows_this, max_items=6)

        except Exception as e:
            inspection_rows.append({"meta.carId": car_id, "meta.error_inspection": str(e)})

        # -----------------
        # ✅ record/open (보험이력)
        # -----------------
        record_json = None
        record_summary = build_record_summary(None)

        if vehicle_no:
            sleepy()
            try:
                record_json = client.fetch_record_open(car_id, vehicle_no)

                rflat = flatten_json(record_json, prefix="record")
                rflat["meta.carId"] = car_id
                rflat["meta.vehicleNo"] = vehicle_no
                record_rows.append(rflat)

                record_accident_rows.extend(explode_record_accidents(car_id, record_json))
                record_summary = build_record_summary(record_json, max_items=3)

            except Exception as e:
                record_rows.append({
                    "meta.carId": car_id,
                    "meta.vehicleNo": vehicle_no,
                    "meta.error_record": str(e)
                })
        else:
            # vehicleNo를 못 찾는 케이스도 있어서 표시만 해둠
            record_rows.append({
                "meta.carId": car_id,
                "meta.vehicleNo": None,
                "meta.error_record": "vehicleNo not found from vehicle response"
            })

        # -----------------
        # ✅ options/choice (옵션)
        # -----------------
        options_summary = build_options_choice_summary(None)
        sleepy()
        try:
            choice = client.fetch_options_choice(car_id)
            options_choice_rows.extend(normalize_options_choice_rows(car_id, choice))
            options_summary = build_options_choice_summary(choice, max_items=15)
        except Exception as e:
            # raw 시트에 에러 형태로 남김
            options_choice_rows.append({"carId": car_id, "meta.error_options_choice": str(e)})

        # -----------------
        # user
        # -----------------
        if user_id:
            if user_id not in user_cache:
                sleepy()
                try:
                    user_cache[user_id] = client.fetch_user(user_id)
                except Exception as e:
                    user_cache[user_id] = {"meta.error_user": str(e)}

            uflat = flatten_json(user_cache[user_id], prefix="user")
            uflat["meta.userId"] = user_id
            uflat["meta.carId"] = car_id
            user_rows.append(uflat)

        # -----------------
        # ✅ summary row (차량당 1행)
        # -----------------
        summary_row = {
            "carId": car_id,
            "userId": user_id,
            "vehicleNo": vehicle_no,

            "inspection.issueCount": issue_count,
            "inspection.issueTop": issue_top,
        }
        summary_row.update(vehicle_summary)

        # record 요약
        summary_row.update(record_summary)

        # options 요약
        summary_row.update(options_summary)

        summary_rows.append(summary_row)

    print("3) Write Excel...")
    wb = Workbook()
    wb.remove(wb.active)

    # ✅ 맨 앞 종합 시트
    write_sheet(wb.create_sheet(), summary_rows, "Summary")

    # 원본/상세
    write_sheet(wb.create_sheet(), list_rows, "list_100")
    write_sheet(wb.create_sheet(), vehicle_rows, "vehicle_detail")
    write_sheet(wb.create_sheet(), inspection_rows, "inspection_detail_raw")
    write_sheet(wb.create_sheet(), user_rows, "seller_user")

    # inspection 세부
    write_sheet(wb.create_sheet(), inspection_item_rows, "inspection_items")
    write_sheet(wb.create_sheet(), inspection_image_rows, "inspection_images")
    write_sheet(wb.create_sheet(), inspection_etc_rows, "inspection_etcs")

    # ✅ 보험이력 세부
    write_sheet(wb.create_sheet(), record_rows, "record_detail_raw")
    write_sheet(wb.create_sheet(), record_accident_rows, "record_accidents")

    # ✅ 옵션 세부
    write_sheet(wb.create_sheet(), options_choice_rows, "options_choice_raw")

    wb.save(OUTPUT_XLSX)
    print(f"✅ DONE: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
